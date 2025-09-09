import openpyxl, re, os, csv
from openpyxl.utils import get_column_letter
from utils import *

from typing import TypeAlias
from openpyxl.worksheet.worksheet import Worksheet

ExtractedData : TypeAlias = tuple[tuple[str, ...], ...]
Patterns : TypeAlias = list[str]

class CSVFileExtractor:
    def __init__(self, col_var):
        self.col_var = col_var

    def export_extracted_data_to_csv(self, output_file : str, patterns : Patterns ,content : str) -> str | None:
        extracted_data = DataExtractor.extract_data(patterns,content)

        if isinstance(extracted_data[0][0], tuple):
            raise TypeError

        if WithLogging.with_logging:
                extracted_data_copy = extracted_data

        if self.col_var.get():
            extracted_data = DataExtractor.create_column_order(extracted_data)

        with open(output_file,'a',newline='',encoding=ENCODING) as f:
            writer = csv.writer(f)
            writer.writerows(extracted_data)

        if WithLogging.with_logging:
            return DataExtractor.get_extracted_data_string(extracted_data_copy)

class ExcelFileExtractor:
    def __init__(self, col_var, exact_var):
        self.col_var = col_var
        self.exact_var = exact_var

    @staticmethod
    def find_max(index : int, sheet : Worksheet) -> int:
        row = 0
        for i in sheet.iter_rows(min_col=index,max_col=index):
            if i[0].value is not None:
                row = i[0].row
        return row

    @staticmethod
    def put_data_in_excel_without_exact_order(extracted_data : ExtractedData, sheet : Worksheet) -> None:
        for data_list in extracted_data:
            sheet.append(data_list)

    @staticmethod
    def get_cell(pattern_letter : str, row_number : int):
        return pattern_letter + str(row_number)

    @staticmethod
    def put_data_in_excel_with_exact_order(extracted_data : ExtractedData, sheet : Worksheet) -> None:
        column_letters_list = [get_column_letter(i) for i in range(1,len(extracted_data)+1)]

        find_max_index = 1
        columns_list_index = 0

        for data_list in extracted_data:
            row_number = ExcelFileExtractor.find_max(find_max_index,sheet) + 1
            for item in data_list:
                sheet[ExcelFileExtractor.get_cell(column_letters_list[columns_list_index],row_number)] = item
                row_number += 1
            columns_list_index += 1
            find_max_index += 1

    def export_extracted_data_to_excel(self, output_file : str, sheet_name : str, patterns : Patterns, content : str) -> str | None:
                if not os.path.isfile(output_file):
                    wb = openpyxl.Workbook()
                    wb.save(output_file)
                    wb.close()

                sheet_name = sheet_name.title()

                wb = openpyxl.load_workbook(output_file)

                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                else:
                    sheet = wb.create_sheet(sheet_name)

                extracted_data = DataExtractor.extract_data(patterns,content)

                if WithLogging.with_logging:
                    extracted_data_copy = extracted_data

                if self.col_var.get() and not self.exact_var.get():
                    extracted_data = DataExtractor.create_column_order(extracted_data)

                if not self.exact_var.get(): # The codes in this if statement will not be executed if 'put in rows' is enabled
                    ExcelFileExtractor.put_data_in_excel_without_exact_order(extracted_data,sheet)
                else:
                    ExcelFileExtractor.put_data_in_excel_with_exact_order(extracted_data,sheet)

                wb.save(output_file)
                wb.close()

                if WithLogging.with_logging:
                    return DataExtractor.get_extracted_data_string(extracted_data_copy)

class DataExtractor:
    def __init__(self, excel_var, log_text, col_var, exact_var):
        self.log_text = log_text
        self.excel_var = excel_var

        self.excel_extractor = ExcelFileExtractor(col_var, exact_var)
        self.csv_extractor = CSVFileExtractor(col_var)

    @staticmethod
    def extract_data(patterns : Patterns, content : str) -> ExtractedData:
        extracted_data = []
        for pattern in patterns:
            data_list = re.findall(pattern,content)
            extracted_data.append(data_list)

        return extracted_data

    def prepare_to_extract_data(self, output_file : str, input_file : str, sheet_name : str, patterns : Patterns) -> None:
            try:
                log_string = ''

                assert patterns, 'There is no patterns to extract data'
                assert output_file, 'The name of output file is required.'

                # --- read the content of input file ---

                with open(input_file,encoding=ENCODING) as f:
                    try:
                        content = f.read()
                    except UnicodeDecodeError:
                        raise ValueError('The input file cannot be a binary file')

                # --- deciding what to do with the output file ---

                output_file_extention = os.path.splitext(output_file)[1].lower()

                if self.excel_var.get():
                    if output_file_extention in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
                        log_string = self.excel_extractor.export_extracted_data_to_excel(output_file,sheet_name,patterns,content)
                    else:
                        raise ValueError('The output file format is not supported. It should be .xlsx, .xlsm, .xltx or .xltm')

                else:
                    log_string = self.csv_extractor.export_extracted_data_to_csv(output_file,patterns,content)

                # --- handling the log ----

                if WithLogging.with_logging:
                    self.log_found_data(log_string, output_file)

            except (FileNotFoundError, AssertionError, PermissionError, ValueError, re.PatternError) as err:
                show_error(err)

            except TypeError as err:
                show_error("You cannot place multiple groups in a pattern")

    @staticmethod
    def get_extracted_data_string(extracted_data_copy : ExtractedData) -> str:
        log_string = ''

        for data_list in extracted_data_copy:
            log_string += '\n'.join(data_list) + '\n'

        return log_string

    def log_found_data(self, log_string : str, output_file : str):
        log_string += f'\n{output_file!r} saved.' + '\n'
        self.log_text.config(state='normal')
        self.log_text.delete('1.0','end')
        self.log_text.insert('end', log_string)
        self.log_text.config(state='disabled')
        self.log_text.see('end')

    @staticmethod
    def create_column_order(extracted_data : ExtractedData) -> tuple[tuple[str]]:
        max_len = max([len(data_list) for data_list in extracted_data])

        for data_list in extracted_data:
            for _ in range(max_len - len(data_list)):
                data_list.append('')

        return tuple(zip(*extracted_data))
