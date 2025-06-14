import openpyxl, re, os, csv
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl.utils import get_column_letter
from tkinter.simpledialog import askstring

from openpyxl.worksheet.worksheet import Worksheet
from typing import TypeAlias, Union

ExtractedData : TypeAlias = tuple[tuple[str, ...], ...]
Patterns : TypeAlias = list[str]

LOG_MODE = ('Disable logging','Enable logging')
APP_TITLE = 'Text2Excel'
ENCODING = 'utf-8-sig'
LOG_DEFAULT = 'log ...'
FILE_TYPES = [
    ('All Files','*.*'),
    ('Excel Files','*.xlsx;*.xlsm;*.xltx;*.xltm'),
    ('CSV Files','*.csv')
]

with_logging = True

MENU_COLOR_ARGS = {'activebackground' : '#00c8ff', 'activeforeground' : 'black'}
EXACT_CB_GRID_ARGS = {'row' : 0, 'column' : 1, 'sticky' : 's', 'pady' : (0,10), 'padx' : (0,20)}

def get_pattern() -> str:
    return askstring(title=APP_TITLE,prompt='Enter the pattern:')

def show_error(err : Union[Exception,str]) -> None:
    messagebox.showerror(title=APP_TITLE, message=err)

def browse_files(widget : ttk.Entry) -> None:
    TITLE = 'Browse'

    if widget == input_file_entry:
        file_path = filedialog.askopenfilename(title=TITLE)
    else:
        file_path = filedialog.askopenfilename(title=TITLE,filetypes=FILE_TYPES)

    if file_path:
        widget.delete(0,'end')
        widget.insert('end',file_path)

class LogMenuCommands:
    def toggle_log() -> None:
        global with_logging
        with_logging = not with_logging
        
        mode = log_menu.entrycget(3,'label')

        if mode == LOG_MODE[0]:
            log_menu.entryconfig(3,label = LOG_MODE[1])
        else:
            log_menu.entryconfig(3,label = LOG_MODE[0])

    def copy_log(event : tk.Event=None) -> None:
        window.clipboard_clear()
        try:
            data = log_text.selection_get()
        except tk.TclError:
            data = log_text.get('1.0','end')
        window.clipboard_append(data)

    def clear_log(event : tk.Event=None) -> None:
        log_text.config(state='normal')
        log_text.delete('1.0','end')
        log_text.insert('end',LOG_DEFAULT)
        log_text.config(state='disabled')

class PatternsMenuCommands:
    def add_pattern(event : tk.Event=None) -> None:
        new_pattern = get_pattern()
        patterns_list.insert('end',new_pattern)

    def insert_pattern(event : tk.Event=None) -> None:
        selected = patterns_list.curselection()
        if len(selected) == 1:
            new_pattern = get_pattern()
            patterns_list.insert(selected[0],new_pattern)

    def edit_selected(event : tk.Event=None) -> None:
        index = patterns_list.curselection()
        if len(index) == 1:
            value = patterns_list.get(index)
            new_value = askstring(title=APP_TITLE, prompt='Enter the pattern: ', initialvalue=value)
            if new_value:
                patterns_list.delete(index)
                patterns_list.insert(index, new_value)

    def delete_selected(event : tk.Event=None) -> None:
        selected = patterns_list.curselection()
        if selected:
            patterns_list.delete(selected[0],selected[-1])
    
    def delete_all(event : tk.Event=None) -> None:
        patterns_list.delete(0,'end')

    def copy_pattern(event : tk.Event=None, all : bool=False) -> None:
        window.clipboard_clear()
        if all:
            window.clipboard_append('\n'.join(patterns_list.get(0,'end')))
        else:
            selected = patterns_list.curselection()
            if selected:
                window.clipboard_append('\n'.join(patterns_list.get(selected[0],selected[-1])))

    def import_from_file(event : tk.Event=None) -> None:
        try:
            file_path = filedialog.askopenfilename(title='Import')
            if file_path:
                with open(file_path, encoding=ENCODING) as f:
                    for i in f.read().strip().splitlines():
                        patterns_list.insert('end', i)
        except UnicodeDecodeError:
            show_error('The patterns file cannot be a binary file')

    def export_to_file(event : tk.Event=None) -> None:
        try:
            file_path = filedialog.asksaveasfilename(title='Export')
            if file_path:
                with open(file_path, 'w',encoding=ENCODING) as f:
                    for i in patterns_list.get(0,'end'):
                        f.write(i + '\n')
        except Exception as err:
            show_error(err)

class CSVFileExtractor:
    def create_csv_file(output_file : str, patterns : Patterns ,content : str) -> str | None:
        extracted_data = DataExtractor.extract_data(patterns,content)

        if with_logging:
                extracted_data_copy = extracted_data

        if col_var.get():
            extracted_data = DataExtractor.create_column_order(extracted_data)

        with open(output_file,'a',newline='',encoding=ENCODING) as f:
            writer = csv.writer(f)
            writer.writerows(extracted_data)
        
        if with_logging:
            return DataExtractor.log_found_data(extracted_data_copy)

class ExcelFileExtractor:
    def find_max(index : int, sheet : Worksheet) -> int:
        row = 0
        for i in sheet.iter_rows(min_col=index,max_col=index):
            if i[0].value is not None:
                row = i[0].row
        return row

    def put_data_in_excel_without_exact_order(extracted_data : ExtractedData, sheet : Worksheet) -> None:
        for data_list in extracted_data:
            sheet.append(data_list)

    def get_cell(pattern_letter : str, row_number : int):
        return pattern_letter + str(row_number)

    def put_data_in_excel_with_exact_order(extracted_data : ExtractedData, sheet : Worksheet) -> None:
        column_letters_list = [get_column_letter(i) for i in range(1,len(extracted_data)+1)]

        find_max_index = 1
        columns_list_index = 0

        for data_list in extracted_data:
            row_number = ExcelFileExtractor.find_max(find_max_index,sheet) + 1
            for item in data_list:
                sheet[ExcelFileExtractor.get_cell(column_letters_list[columns_list_index],row_number)] = item
                row_number += 1
            columns_list_index += 1
            find_max_index += 1

    def create_excel_file(output_file : str, sheet_name : str, patterns : Patterns, content : str) -> str | None:
                if not os.path.isfile(output_file):
                    wb = openpyxl.Workbook()
                    wb.save(output_file)
                    wb.close()
                
                sheet_name = sheet_name.title()

                wb = openpyxl.load_workbook(output_file)

                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                else:
                    sheet = wb.create_sheet(sheet_name)

                extracted_data = DataExtractor.extract_data(patterns,content)

                if with_logging:
                    extracted_data_copy = extracted_data

                if col_var.get() and not exact_var.get():
                    extracted_data = DataExtractor.create_column_order(extracted_data)
                
                if not exact_var.get(): # The codes in this if statement will not be executed if 'put in rows' is enabled
                    ExcelFileExtractor.put_data_in_excel_without_exact_order(extracted_data,sheet)
                else:
                    ExcelFileExtractor.put_data_in_excel_with_exact_order(extracted_data,sheet)

                wb.save(output_file)
                wb.close()

                if with_logging:
                    return DataExtractor.log_found_data(extracted_data_copy)

class DataExtractor:
    def extract_data(patterns : Patterns, content : str) -> ExtractedData:
        extracted_data = []
        for pattern in patterns:
            data_list = re.findall(pattern,content)
            extracted_data.append(data_list)

        return extracted_data

    def prepare_to_extract_data(output_file : str, input_file : str, sheet_name : str, patterns : Patterns) -> None:
            try:
                assert patterns, 'There is no patterns to extract data'

                with open(input_file,encoding=ENCODING) as f:
                    try:
                        content = f.read()
                    except UnicodeDecodeError:
                        raise ValueError('The input file cannot be a binary file')

                assert output_file, 'The name of output file is required.'
                
                output_file_extention = os.path.splitext(output_file)[1].lower()

                if excel_var.get():
                    if output_file_extention in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
                        log_string = ExcelFileExtractor.create_excel_file(output_file,sheet_name,patterns,content)
                    else:
                        raise ValueError('The output file format is not supported. It should be .xlsx, .xlsm, .xltx or .xltm')

                else:
                    log_string = CSVFileExtractor.create_csv_file(output_file,patterns,content)

                if with_logging:
                    log_string += f'\n{output_file!r} saved.' + '\n'
                    log_text.config(state='normal')
                    log_text.delete('1.0','end')
                    log_text.insert('end', log_string)
                    log_text.config(state='disabled')
                    log_text.see('end')

            except (FileNotFoundError, AssertionError, PermissionError, ValueError) as err:
                show_error(err)

            
    def log_found_data(extracted_data_copy : ExtractedData) -> str:
        log_string = ''

        for data_list in extracted_data_copy:
            log_string += '\n'.join(data_list) + '\n'

        return log_string
    
    def create_column_order(extracted_data : ExtractedData) -> tuple[tuple[str]]:
        max_len = max([len(data_list) for data_list in extracted_data])

        for data_list in extracted_data:
            for _ in range(max_len - len(data_list)):
                data_list.append('')
            
        return tuple(zip(*extracted_data))

class ContextMenuDisplayers:
    def show_log_menu(event : tk.Event, app : bool = False) -> None:
        if log_text.tag_ranges('sel'):
            text='Copy selected'
        else:
            text='Copy log'
        log_menu.entryconfig(0,label=text)

        if app:
            log_menu.tk_popup(log_text.winfo_rootx()+100,log_text.winfo_rooty()+100)
        else:
            log_menu.tk_popup(event.x_root,event.y_root)
            

    def show_patterns_menu(event : tk.Event, app : bool=False) -> None:
        selected = patterns_list.curselection()
        if selected:
            if len(selected)>1:
                for i in (1,3):
                    patterns_menu.entryconfig(i,state='disabled')
                
                for i in range(4,6):
                    patterns_menu.entryconfig(i,state='active')

            else:
                patterns_menu.entryconfig(1,state='active')
                for i in range(3,6):
                    patterns_menu.entryconfig(i,state='active')
        else:
            patterns_menu.entryconfig(1,state='disabled')
            for i in range(3,6):
                patterns_menu.entryconfig(i,state='disabled')

        if app:
            patterns_menu.tk_popup(patterns_list.winfo_rootx()+100,patterns_list.winfo_rooty()+100)
        else:
            patterns_menu.tk_popup(event.x_root, event.y_root)

    def show_entry_menu(menu : tk.Menu, event : tk.Event, app : bool=False) -> None:
        if window.focus_get() == event.widget:
            for i in range(4):
                menu.entryconfig(i,state='active')
        else:
            for i in range(4):
                menu.entryconfig(i,state='disabled')
        if app:
            if event.widget == sheet_name_entry:
                menu.tk_popup(event.widget.winfo_rootx()+50,event.widget.winfo_rooty()+25)
            else:
                menu.tk_popup(event.widget.winfo_rootx()+100,event.widget.winfo_rooty()+25)
        else:
            menu.tk_popup(event.x_root,event.y_root)

class MenuCreators:
    def create_patterns_menu() -> tk.Menu:
        menu = tk.Menu(tearoff=False,**MENU_COLOR_ARGS)
        menu.add_command(label='Add Pattern', command=PatternsMenuCommands.add_pattern,accelerator='Ctrl+Shift+A')
        menu.add_command(label='Insert Pattern',command=PatternsMenuCommands.insert_pattern,accelerator='Ctrl+I')
        menu.add_separator()
        menu.add_command(label='Edit selected', command=PatternsMenuCommands.edit_selected,accelerator='F2')
        menu.add_command(label='Delete selected', command=PatternsMenuCommands.delete_selected,accelerator='Delete')
        menu.add_command(label='Copy selected', command=PatternsMenuCommands.copy_pattern,accelerator='Ctrl+C')
        menu.add_command(label='Delete All', command=PatternsMenuCommands.delete_all,accelerator='Ctrl+Shift+D')
        menu.add_command(label='Copy All', command=lambda : PatternsMenuCommands.copy_pattern(all=True),accelerator='Ctrl+Shift+C')
        menu.add_separator()
        menu.add_command(label='Import from file', command=PatternsMenuCommands.import_from_file,accelerator='Ctrl+Shift+I')
        menu.add_command(label='Export to file', command=PatternsMenuCommands.export_to_file,accelerator='Ctrl+E')
        return menu

    def create_log_menu() -> tk.Menu:
        menu = tk.Menu(tearoff=False,**MENU_COLOR_ARGS)
        menu.add_command(label='Copy log',command=LogMenuCommands.copy_log,accelerator='Ctrl+C')
        menu.add_command(label='Clear log',command=LogMenuCommands.clear_log,accelerator='Ctrl+D')
        menu.add_separator()
        menu.add_command(label=LOG_MODE[0],command=LogMenuCommands.toggle_log)
        return menu

    def create_entry_menu(widget : ttk.Entry, is_file_entry : bool=True,is_output_file_entry : bool=True) -> tk.Menu:
        menu = tk.Menu(tearoff=False,**MENU_COLOR_ARGS)
        menu.add_command(label='Select All', accelerator='Ctrl+A',command=lambda : widget.select_range(0,'end'))
        menu.add_command(label='Copy', accelerator='Ctrl+C',command=lambda : widget.event_generate('<<Copy>>'))
        menu.add_command(label='Paste', accelerator='Ctrl+V',command=lambda : widget.event_generate('<<Paste>>'))
        menu.add_command(label='Cut', accelerator='Ctrl+X',command=lambda : widget.event_generate('<<Cut>>'))
        menu.add_separator()
        menu.add_command(label='Clear',accelerator='Ctrl+Shift+C',command=lambda : widget.delete(0,'end'))
        if is_file_entry:
            menu.add_command(label='Browse',command=lambda : browse_files(widget),accelerator='Ctrl+B')
            if is_output_file_entry:
                menu.add_separator()
                menu.add_radiobutton(label='Excel',variable=excel_var,value=True,command=CSVExcelSwitchFunctions.show_only_excel_required_widgets)
                menu.add_radiobutton(label='CSV',variable=excel_var,value=False,command=CSVExcelSwitchFunctions.hide_only_excel_required_widgets)
        return menu

class CSVExcelSwitchFunctions:

    exact_var_value = None
    def hide_exact_order_cb() -> None:
            exact_cb.grid_remove()
            CSVExcelSwitchFunctions.exact_var_value = exact_var.get()
            exact_var.set(False)
            exact_cb_substitute_lbl.grid(**EXACT_CB_GRID_ARGS)

    def show_exact_order_cb() -> None:
        if excel_var.get():
            exact_cb_substitute_lbl.grid_remove()
            exact_var.set(CSVExcelSwitchFunctions.exact_var_value)
            exact_cb.grid()

    def hide_only_excel_required_widgets() -> None: # sheet_name_entry, sheet_name_lbl, exact_cb
        if exact_cb.winfo_ismapped():
            CSVExcelSwitchFunctions.hide_exact_order_cb()
        
        sheet_name_lbl.grid_remove()
        sheet_name_entry.grid_remove()

    def show_only_excel_required_widgets() -> None:
        if col_var.get():
            CSVExcelSwitchFunctions.show_exact_order_cb()

        sheet_name_lbl.grid()
        sheet_name_entry.grid()

window = tk.Tk()
window.title(APP_TITLE)
window.resizable(False,False)

icon_path = os.path.join(os.path.dirname(__file__),'icon.png')
icon_photo = tk.PhotoImage(file=icon_path)
window.iconphoto(True,icon_photo)

main_frm = tk.Frame()
main_frm.grid(row=0,column=0, rowspan=3,sticky='nsew')

frm = tk.Frame(main_frm,borderwidth=10)
frm.pack(fill='both',expand=1)

input_file_lbl = tk.Label(frm,text='Input file:')
output_file_lbl = tk.Label(frm,text='Output file:')
sheet_name_lbl = tk.Label(frm,text='Sheet name:')

input_file_entry = ttk.Entry(frm,width=30)
output_file_entry = ttk.Entry(frm,width=30)
sheet_name_entry = ttk.Entry(frm,width=15)

input_file_lbl.grid(row=2,column=0,sticky='w')
output_file_lbl.grid(row=2, column=3,sticky='w')
sheet_name_lbl.grid(row=4, column=3, sticky='w')

input_file_entry.grid(row=3,column=0,sticky='w')
output_file_entry.grid(row=3,column=3,sticky='w')
sheet_name_entry.grid(row=5, column=3, sticky='w')

log_frm = tk.Frame(main_frm)
yscroll_log = tk.Scrollbar(log_frm)
xscroll_log = tk.Scrollbar(log_frm, orient='horizontal')
log_text = tk.Text(log_frm,width=23, height=10, font = 'TkTextFont', wrap = 'none', 
                   yscrollcommand=yscroll_log.set, xscrollcommand=xscroll_log.set,takefocus=True,
                   highlightcolor='black',highlightthickness=1)
yscroll_log.config(command=log_text.yview)
xscroll_log.config(command=log_text.xview)
log_text.insert('end', LOG_DEFAULT)
log_text.config(state='disabled')
log_text.grid(row=0,column=0)
xscroll_log.grid(row=1,column=0, pady=(0,10), sticky='we')
yscroll_log.grid(row=0,column=1, sticky='ns')
log_frm.pack()

exact_cb_substitute_lbl = tk.Label()
exact_var = tk.IntVar()
exact_cb = ttk.Checkbutton(text='Exact order', variable=exact_var)
exact_cb.grid(**EXACT_CB_GRID_ARGS)

column_row_frm = tk.Frame()
col_var = tk.IntVar(value=True)
col_rb = ttk.Radiobutton(column_row_frm,text='Put in columns', variable=col_var, value=True,command=CSVExcelSwitchFunctions.show_exact_order_cb)
row_rb = ttk.Radiobutton(column_row_frm,text='Put in rows', variable=col_var, value=False,command=CSVExcelSwitchFunctions.hide_exact_order_cb)
col_rb.pack(anchor='w')
row_rb.pack(anchor='w')
column_row_frm.grid(row=1,column=1, sticky='s')

excel_var = tk.IntVar(value=True)

patterns_list_frm = tk.Frame(bd = 10)
yscroll_pl = tk.Scrollbar(patterns_list_frm)
xscroll_pl = tk.Scrollbar(patterns_list_frm, orient='horizontal')
pattern_lbl = tk.Label(patterns_list_frm,text='Patterns:')
patterns_list = tk.Listbox(patterns_list_frm,width=25,height=13, yscrollcommand=yscroll_pl.set, 
                           xscrollcommand=xscroll_pl.set,selectmode='extended')
xscroll_pl.config(command=patterns_list.xview)
yscroll_pl.config(command=patterns_list.yview)
pattern_lbl.grid(row=0,column=0, sticky='w')
patterns_list.grid(row=1,column=0)
xscroll_pl.grid(row=2,column=0, sticky='we')
yscroll_pl.grid(row=1,column=1,sticky='ns')
patterns_list_frm.grid(row=2,column=1, sticky='s')

input_file_menu = MenuCreators.create_entry_menu(input_file_entry,is_output_file_entry=False)
output_file_menu = MenuCreators.create_entry_menu(output_file_entry)
sheet_name_menu = MenuCreators.create_entry_menu(sheet_name_entry,False)
patterns_menu  = MenuCreators.create_patterns_menu()
log_menu = MenuCreators.create_log_menu()

patterns_list.bind('<Button-3>', ContextMenuDisplayers.show_patterns_menu)
patterns_list.bind('<App>', lambda event : ContextMenuDisplayers.show_patterns_menu(event, True))

log_text.bind('<Button-3>',lambda event : ContextMenuDisplayers.show_log_menu(event))
log_text.bind('<App>',lambda event : ContextMenuDisplayers.show_log_menu(event,True))

input_file_entry.bind('<Button-3>',lambda event : ContextMenuDisplayers.show_entry_menu(input_file_menu,event))
input_file_entry.bind('<App>',lambda event : ContextMenuDisplayers.show_entry_menu(input_file_menu,event,True))

output_file_entry.bind('<Button-3>',lambda event : ContextMenuDisplayers.show_entry_menu(output_file_menu,event))
output_file_entry.bind('<App>',lambda event : ContextMenuDisplayers.show_entry_menu(output_file_menu,event,True))

sheet_name_entry.bind('<Button-3>', lambda event : ContextMenuDisplayers.show_entry_menu(sheet_name_menu,event))
sheet_name_entry.bind('<App>',lambda event : ContextMenuDisplayers.show_entry_menu(sheet_name_menu,event,True))

window.bind_class('TEntry','<Control-C>',lambda event : event.widget.delete(0,'end'))

log_text.bind('<Control-c>',LogMenuCommands.copy_log)
log_text.bind('<Control-d>',LogMenuCommands.clear_log)

input_file_entry.bind('<Control-b>',lambda event  : browse_files(event.widget))
output_file_entry.bind('<Control-b>',lambda event : browse_files(event.widget))

input_file_entry.focus_set()

log_text.bind('<FocusOut>',lambda event : log_text.tag_remove('sel','1.0','end'))

patterns_list.bind('<Control-A>',PatternsMenuCommands.add_pattern)
patterns_list.bind('<Control-i>',PatternsMenuCommands.insert_pattern)
patterns_list.bind('<F2>',PatternsMenuCommands.edit_selected)
patterns_list.bind('<Delete>',PatternsMenuCommands.delete_selected)
patterns_list.bind('<Control-d>',PatternsMenuCommands.delete_selected)
patterns_list.bind('<Control-c>',PatternsMenuCommands.copy_pattern)
patterns_list.bind('<Control-D>',PatternsMenuCommands.delete_all)
patterns_list.bind('<Control-C>',lambda event : PatternsMenuCommands.copy_pattern(all=True))
patterns_list.bind('<Control-I>',PatternsMenuCommands.import_from_file)
patterns_list.bind('<Control-e>',PatternsMenuCommands.export_to_file)

btn_convert = tk.Button(frm,text='convert',width=10,height=5,background='#0080e5',
                        command=lambda : DataExtractor.prepare_to_extract_data(output_file_entry.get(), input_file_entry.get(), sheet_name_entry.get(), 
                        patterns_list.get(0,'end'))
                        ,cursor='hand2')

btn_convert.bind('<Enter>', lambda event : btn_convert.config(bg = '#0092ff'))
btn_convert.bind('<Leave>', lambda event : btn_convert.config(bg = '#0080e5'))
btn_convert.grid(row=0,column=1)

window.mainloop()
